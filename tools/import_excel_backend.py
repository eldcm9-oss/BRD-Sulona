from __future__ import annotations

import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "backend" / "data" / "ims-db.json"


def scalar(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def table_rows(ws, header_row=4):
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = [scalar(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def number(value, default=0):
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def project_health(status, progress, finish):
    if status == "Overdue":
        return "Red"
    if status == "Completed":
        return "Green"
    if number(progress) < 50:
        return "Yellow"
    return "Green"


def percent_0_to_100(value):
    parsed = number(value)
    if 0 < parsed <= 1:
        return round(parsed * 100, 2)
    return parsed


def normalize_projects(rows):
    projects = []
    for row in rows:
        project_id = row.get("Project ID")
        brd = row.get("BRD Number")
        name = row.get("Project Name")
        if not project_id or not name:
            continue
        status = row.get("Status") or "Draft"
        progress = percent_0_to_100(row.get("Progress %"))
        projects.append(
            {
                "id": str(project_id),
                "brd": str(brd or project_id),
                "name": str(name),
                "division": row.get("Division") or "",
                "department": row.get("Department") or "",
                "site": row.get("Site") or "",
                "category": row.get("Category") or "",
                "priority": row.get("Priority") or "Medium",
                "status": status,
                "health": project_health(status, progress, row.get("Planned Finish Date")),
                "pic": row.get("PIC") or "",
                "vendor": row.get("Vendor") or "",
                "budget": number(row.get("Budget Requested")),
                "approved": number(row.get("Budget Approved")),
                "actual": number(row.get("Actual Spend")),
                "progress": progress,
                "start": row.get("Planned Start Date") or "",
                "finish": row.get("Planned Finish Date") or "",
                "approvalDays": 0,
                "objective": row.get("Project Objective") or "",
                "remarks": row.get("Remarks") or "",
            }
        )
    return projects


def normalize_issues(rows, project_name_by_id):
    issues = []
    for row in rows:
        issue_id = row.get("Issue ID")
        title = row.get("Title")
        if not issue_id or not title:
            continue
        project_id = row.get("Project ID") or ""
        issues.append(
            {
                "id": str(issue_id),
                "projectId": str(project_id),
                "project": project_name_by_id.get(project_id, str(project_id)),
                "title": str(title),
                "category": row.get("Category") or "",
                "severity": row.get("Severity") or "Medium",
                "priority": row.get("Priority") or "Medium",
                "due": row.get("Due Date") or "",
                "owner": row.get("Owner") or "",
                "status": row.get("Status") or "Open",
                "resolution": row.get("Resolution") or "",
            }
        )
    return issues


def normalize_activities(rows, project_name_by_id):
    activities = []
    for row in rows:
        activity_id = row.get("Activity ID")
        activity_type = row.get("Activity Type")
        if not activity_id or not activity_type:
            continue
        project_id = row.get("Project ID") or ""
        activities.append(
            {
                "id": str(activity_id),
                "projectId": str(project_id),
                "date": row.get("Activity Date") or "",
                "type": str(activity_type),
                "project": project_name_by_id.get(project_id, str(project_id)),
                "text": row.get("Description") or "",
                "pic": row.get("PIC") or "",
                "status": row.get("Status") or "",
            }
        )
    return activities


def normalize_collection(rows, id_key):
    normalized = []
    for row in rows:
        row_id = row.get(id_key)
        if not row_id:
            continue
        item = {str(key): scalar(value) for key, value in row.items() if key}
        item["id"] = str(row_id)
        normalized.append(item)
    return normalized


def master_data(ws):
    headers = [ws.cell(4, col).value for col in range(1, ws.max_column + 1)]
    mapping = {
        "Division": "divisions",
        "Department": "departments",
        "Site": "sites",
        "Vendor": "vendors",
        "Project Category": "categories",
        "Project Status": "statuses",
        "Priority": "priorities",
        "Issue Category": "issueCategories",
        "Issue Severity": "issueSeverities",
        "Document Category": "documentCategories",
    }
    result = {value: [] for value in mapping.values()}
    for col, header in enumerate(headers, start=1):
        key = mapping.get(header)
        if not key:
            continue
        for row in range(5, ws.max_row + 1):
            value = ws.cell(row, col).value
            if value not in (None, ""):
                result[key].append(str(value))
    return result


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: import_excel_backend.py <workbook.xlsx>")

    workbook_path = Path(sys.argv[1])
    wb = load_workbook(workbook_path, data_only=True)

    projects = normalize_projects(table_rows(wb["Projects"]))
    project_name_by_id = {project["id"]: project["name"] for project in projects}

    db = {
        "meta": {
            "sourceWorkbook": str(workbook_path),
            "importedAt": datetime.now().isoformat(timespec="seconds"),
            "nextProjectNumber": 26013,
        },
        "projects": projects,
        "issues": normalize_issues(table_rows(wb["Issues"]), project_name_by_id),
        "activities": normalize_activities(table_rows(wb["Activities"]), project_name_by_id),
        "approvals": normalize_collection(table_rows(wb["Approvals"]), "Approval ID"),
        "procurements": normalize_collection(table_rows(wb["Procurement"]), "Procurement ID"),
        "documents": normalize_collection(table_rows(wb["Documents"]), "Document ID"),
        "reviews": normalize_collection(table_rows(wb["Reviews"]), "Review ID"),
        "masterData": master_data(wb["Master Data"]),
    }

    if DB_PATH.exists():
        backup = DB_PATH.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d%H%M%S')}.json")
        shutil.copy2(DB_PATH, backup)

    DB_PATH.write_text(json.dumps(db, indent=2), encoding="utf-8")
    print(json.dumps({
        "dbPath": str(DB_PATH),
        "projects": len(db["projects"]),
        "issues": len(db["issues"]),
        "activities": len(db["activities"]),
        "approvals": len(db["approvals"]),
        "documents": len(db["documents"]),
    }, indent=2))


if __name__ == "__main__":
    main()
